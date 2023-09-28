import json
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from R4C.settings import BASE_DIR
from robots.models import Robot


@method_decorator(csrf_exempt, name='dispatch')
class RobotsFactory(View):

    def post(self, request):
        data = json.loads(request.body.decode('utf-8'))
        robot_model = str(data.get('model'))
        robot_version = str(data.get('version'))
        robot_serial = '{0}-{1}'.format(robot_model, robot_version)
        robot_created = data.get('created')
        robot_data = {
            'serial': robot_serial.lower(),
            'model': robot_model.lower(),
            'version': robot_version.lower(),
            'created': robot_created,
        }
        robot_item = Robot.objects.create(**robot_data)
        message = 'New robot added to warehouse with id: {0}'.format(
            robot_item.id,
        )
        response_data = {"message": message}
        return JsonResponse(response_data, status=201)


class FactoryReport(View):
    def get(self, request):
        with open(os.path.join(BASE_DIR, 'models_list.json')) as json_file:
            models_list = json.load(json_file)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet',
        )
        content_disp = 'attachment; filename={0}-factory_report.xlsx'.format(
            datetime.now().strftime('%Y-%m-%d'),
        )
        response['Content-Disposition'] = content_disp

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='center')
        border_bottom = Border(
            bottom=Side(border_style='medium', color='FF000000'),
        )
        wrapped_alignment = Alignment(vertical='top', wrap_text=True)
        columns = [('Model', 35), ('Version', 35), ('Count', 35)]
        for model_index, model in enumerate(models_list.get('robot_models')):
            worksheet = workbook.create_sheet(
                title=model,
                index=model_index,
            )
            fill = PatternFill(fill_type='solid')
            row_num = 1
            for col_num, (column_title, column_width) in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.border = border_bottom
                cell.alignment = centered_alignment
                cell.fill = fill
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = column_width
            first_date = datetime.now() - timedelta(days=7)
            robots = Robot.objects.filter(
                model=model,
                created__gt=first_date,
            ).order_by('created')
            robot_version = []
            for robot in robots:
                robot_version.append(robot.version)
            for version in set(robot_version):
                row_num += 1
                row = [
                    (model, 'Normal'),
                    (version, 'Normal'),
                    (robots.filter(version=version).count(), 'Normal'),
                ]
                for col_num, (cell_value, cell_format) in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.style = cell_format
                    cell.alignment = wrapped_alignment
            worksheet.freeze_panes = worksheet['A2']
        workbook.save(response)
        return response
